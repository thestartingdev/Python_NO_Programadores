# Vamos a llamar a la libreria openpyxl 
# que trabaja con hojas de calculo
# previo la instalamos con
# pip install openpyxl
import openpyxl

# vamos a abrir una hoja de calculo llamada
# example.xlsx
wb = openpyxl.load_workbook('example.xlsx')
type(wb)

# Ahora vamos a traer los nombres de las hojas de calculo
wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames

# Nos traemos el nombre de una hoja solamente
sheet = wb['Sheet3']
sheet

# Nos traemos el titulo
sheet.title